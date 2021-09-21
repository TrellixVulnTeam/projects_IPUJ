import openpyxl

path = "D:\sample.xlsx"
wb = openpyxl.load_workbook(path)
sheet1 = wb["Sheet1"]
sheet2 = wb["Sheet2"]
sheet3 = wb["Sheet3"]

rows = sheet3.max_row
columns = sheet3.max_column
print("rows : ", rows)
print("columns :", columns)

#To Extract data from excel
for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet3.cell(row = r, column = c).value, end=" ")
    print()

# # To load data into excel
# for r in range(1, 7):
#     for c in range(1, 4):
#         sheet2.cell(row = r, column = c).value="hello"
# wb.save(path)

