import openpyxl

def get_row_count(file_name, sheet_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    return(sheet.max_row)

def get_column_count(file_name, sheet_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    return (sheet.max_column)

def read_data(file_name, sheet_name, row_num, column_num):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    return (sheet.cell(row = row_num, column = column_num).value)

def write_data(file_name, sheet_name, row_num, column_num, data):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    sheet.cell(row=row_num, column=column_num).value = data
    wb.save(file_name)


